"""Parse CSV / Excel uploads into plain-text lines for agent_catalog (pipe-separated cells)."""

from __future__ import annotations

import csv
import io
import re
from typing import Any

from openpyxl import load_workbook

MAX_BYTES = 2 * 1024 * 1024
MAX_ROWS = 3000


def _decode_text(data: bytes) -> str:
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError:
        return data.decode("latin-1", errors="replace")


def _row_to_line(cells: list[Any]) -> str | None:
    parts: list[str] = []
    for c in cells:
        if c is None:
            s = ""
        else:
            s = str(c).strip().replace("\n", " ").replace("\r", "")
        parts.append(s)
    if not any(parts):
        return None
    return " | ".join(parts)


def _rows_to_text(rows: list[list[Any]]) -> tuple[str, int]:
    lines: list[str] = []
    for row in rows[:MAX_ROWS]:
        line = _row_to_line(row)
        if line is not None:
            lines.append(line)
    return "\n".join(lines), len(lines)


def parse_csv(data: bytes) -> tuple[str, int]:
    text = _decode_text(data)
    sample = text[:8192]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        if sample.count(";") > sample.count(","):
            delimiter = ";"
        elif "\t" in sample[:200]:
            delimiter = "\t"
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows: list[list[Any]] = []
    for i, row in enumerate(reader):
        if i >= MAX_ROWS:
            break
        rows.append(list(row))
    return _rows_to_text(rows)


def parse_xlsx(data: bytes) -> tuple[str, int]:
    bio = io.BytesIO(data)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: list[list[Any]] = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_ROWS:
                break
            rows.append(list(r))
        return _rows_to_text(rows)
    finally:
        wb.close()


def parse_catalog_bytes(data: bytes, filename: str) -> tuple[str, int, str]:
    if len(data) > MAX_BYTES:
        raise ValueError(f"El archivo supera {MAX_BYTES // (1024 * 1024)} MB.")
    name = (filename or "").lower()
    ext = name.rsplit(".", 1)[-1] if "." in name else ""
    ext = re.sub(r"[^a-z0-9]", "", ext)
    if ext in ("csv", "txt"):
        text, n = parse_csv(data)
        kind = "csv"
    elif ext == "xlsx":
        text, n = parse_xlsx(data)
        kind = "xlsx"
    else:
        raise ValueError("Formato no soportado. Usa .csv, .txt (CSV) o .xlsx (Excel).")
    if not (text or "").strip():
        raise ValueError("No se encontraron filas con datos legibles en el archivo.")
    return text.strip(), n, kind
